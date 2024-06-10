from openpyxl import Workbook
import os
import datetime
import pandas as pd

class PokemonDataProcessor:
    def __init__(self, input_file, locations_file, output_file, ignore_list, key_order, moves_tab, poke_info_tab, misc_tab):
        self.input_file = input_file
        self.locations_file = locations_file
        self.output_file = output_file
        self.ignore_list = ignore_list
        self.key_order = key_order
        self.moves_tab = moves_tab
        self.poke_info_tab = poke_info_tab
        self.misc_tab = misc_tab
        self.all_data = []
        self.location_data = {}

    # Split and process pokemon base file into Python data dictionary    
    def process_input_file(self):
        with open(self.input_file, 'r') as file:
            lines = file.readlines()

        data = {}
        for line in lines:
            line = line.strip()
            if line == '#-------------------------------':
                if data:
                    if 'InternalName' not in data or data['InternalName'] is None:
                        data['InternalName'] = data['Name'].upper()
                    data['UniqueID'] = f"{data['Name']}_{data['InternalName']}"  # Generate UniqueID
                    self.all_data.append(data)
                    data = {}
            elif '=' in line:
                key, value = line.split('=', 1)
                key = key.strip()
                value = value.strip()
                # print(key)
                if key in self.ignore_list and not 'Types':
                    continue
                if key == 'BaseStats':
                    # print('BaseStats')
                    stats = value.split(',')
                    data['HP'] = stats[0]
                    data['Attack'] = stats[1]
                    data['Defense'] = stats[2]
                    data['Spd'] = stats[3]
                    data['SpAtk'] = stats[4]
                    data['SpDef'] = stats[5]
                if key == 'HiddenAbility':
                    # print('HiddenAbility')
                    abilities = value.split(',')
                    for ability in abilities:
                        data['HiddenAbility'] = ability
                if key == 'Types':
                    # print('Types')
                    types = value.split(',')
                    data['Type1'] = types[0]
                    if len(types) > 1:
                        data['Type2'] = types[1]
                    else:
                        data['Type2'] = None
                else:
                    data[key] = value

        if data:
            if 'InternalName' not in data or data['InternalName'] is None:
                data['InternalName'] = data['Name'].upper()
            data['UniqueID'] = f"{data['Name']}_{data['InternalName']}"

            # Check if InternalName is EEVEE and update Evolutions field
            if data.get('InternalName') == 'EEVEE':
                if 'Evolutions' in data:
                    data['Evolutions'] += '(32)'
                else:
                    data['Evolutions'] = '(32)'

            self.all_data.append(data)


    # Split and process encounters file into Python data dictionary 
    def process_locations_file(self):
        with open(self.locations_file, 'r') as file:
            lines = file.readlines()
        
        current_location = None
        for line in lines:
            line = line.strip()
            if line.startswith('#'):
                continue
            if line.startswith('['):
                current_location = line.split('#')[1].strip()
            elif ',' in line:
                parts = line.split(',')
                if len(parts) == 4:
                    _, name, level_start, level_end = parts
                    level_range = f"{level_start}-{level_end}"
                    location_entry = f"{current_location}: {level_range}"
                    if name not in self.location_data:
                        self.location_data[name] = []
                    self.location_data[name].append(location_entry)

    # Add location data to pokemon names
    def add_location_data(self):
        for entry in self.all_data:
            name = entry.get('InternalName', '')
            if name in self.location_data:
                entry['Location Found'] = "; ".join(self.location_data[name])
            else:
                entry['Location Found'] = ''
        
        if 'Location Found' not in self.key_order:
            self.key_order.append('Location Found')

    # Build Excel workbook
    def create_workbook(self):
        wb = Workbook()
        
        def add_data_to_sheet(sheet_name, data, headers):
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)
            for entry in data:
                row = [entry.get(header, '') for header in headers]
                ws.append(row)
        
        # Build the different sheets based on orders below
        main_headers = ['InternalName', 'Name'] + self.key_order if 'InternalName' not in self.key_order else self.key_order
        add_data_to_sheet("Main", self.all_data, main_headers)
        print('adding to main')
        add_data_to_sheet("Moves", self.all_data, ['InternalName', 'Name'] + self.moves_tab)
        add_data_to_sheet("Poke Info", self.all_data, ['InternalName', 'Name'] + self.poke_info_tab)
        add_data_to_sheet("Misc", self.all_data, ['InternalName', 'Name'] + self.misc_tab)
        
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        wb.save(self.output_file)

    # Check if the Excel sheet exists and has all required columns
    def check_excel_sheet(self):
        if not os.path.exists(self.output_file):
            return False
        
        required_columns = ['Name', 'InternalName', 'FormName', 'Type1', 'Type2', 'HP', 'Attack', 'Defense', 'SpAtk', 'SpDef', 'Spd', 'Abilities', 'HiddenAbility', 'Evolutions', 'Location Found', 'Evolution Line']
        
        try:
            df = pd.read_excel(self.output_file, sheet_name="Main")
            columns = df.columns.tolist()
            print(columns)
            
            if set(required_columns).issubset(set(columns)):
                return True
            else:
                print('fail')
                return False
                
        except Exception as e:
            #print(f"An error occurred while checking Excel sheet: {e}")
            return False
    
    # Check if the Excel sheet was created in the last 2 days
    def check_excel_creation_date(self):
        if os.path.exists(self.output_file):
            creation_time = datetime.datetime.fromtimestamp(os.path.getctime(self.output_file))
            current_time = datetime.datetime.now()
            if (current_time - creation_time).days <= 2:
                print('pass')
                return True
        return False

    # Run all functions
    def run(self):
        if self.check_excel_sheet() and self.check_excel_creation_date():
            pass
        else:
            self.process_input_file()
            self.process_locations_file()
            self.add_location_data()
            self.create_workbook()

# Get the directory of the current script file
#base_dir = os.path.dirname(os.path.abspath(__file__))

# For WORKING_PATH
base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
# For pyinstaller
#base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Combine the directory path and the file name
workbook_path = os.path.join(base_dir, 'PBS', 'pokemon.txt')
encounter_path = os.path.join(base_dir, 'PBS', 'encounters.txt')

# Configuration
input_file = workbook_path
locations_file = encounter_path
output_file = 'pokedexCEL.xlsx'
ignore_list = ['BaseExp', 'BattlerEnemyX', 'BattlerEnemyY', 'BattlerPlayerX', 'BattlerPlayerY', 'BattlerShadowSize', 'BattlerShadowX', 'CatchRate', 'Types', 'Category', 'EggGroups', 'HatchSteps', 'EVs', 'EffortPoints', 'HiddenAbilities']
key_order = ['UniqueID', 'Name', 'InternalName', 'FormName', 'Type1', 'Type2', 'HP', 'Attack', 'Defense', 'SpAtk', 'SpDef', 'Spd', 'Abilities', 'HiddenAbility', 'Evolutions', 'Location Found']
moves_tab = ['UniqueID', 'InternalName', 'Moves', 'TutorMoves', 'EggMoves']
poke_info_tab = ['UniqueID', 'Generation', 'Height', 'Weight', 'Color', 'Pokedex', 'Shape', 'Kind', 'GrowthRate', 'GenderRate', 'GenderRatio', 'Habitat', 'BaseEXP', 'Rareness', 'Happiness', 'Compatibility', 'Incense']
misc_tab = ['WildItemCommon', 'WildItemRare', 'WildItemUncommon', 'StepsToHatch']

# Run the processor
processor = PokemonDataProcessor(input_file, locations_file, output_file, ignore_list, key_order, moves_tab, poke_info_tab, misc_tab)

def main():
    processor.run()

# REMOVE WHEN PYINSTALLERING
#main()