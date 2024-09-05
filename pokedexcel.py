import re
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
import datetime

class MoveParser:
    def __init__(self, filename):
        self.filename = filename
        self.moves = self.parse_moves()
    
    import re

    def parse_moves(self):
        with open(self.filename, 'r') as file:
            content = file.read()

        moves = {}
        move_blocks = content.split('#-------------------------------')
        move_pattern = re.compile(r"\[(.*?)\](.*?)Name = (.*?)\nType = (.*?)\nCategory = (.*?)\n(?:Power = (.*?)\n)?Accuracy = (.*?)\n", re.S)

        for block in move_blocks:
            match = move_pattern.search(block)
            if match:
                move_code, _, name, move_type, category, power, accuracy = match.groups()
                # If power is None (i.e., not captured), set it to 'N/A'
                power = power.strip() if power else 'N/A'
                moves[move_code.strip().upper()] = {
                    'Name': name.strip(),
                    'Type': move_type.strip(),
                    'Category': category.strip(),
                    'Power': power,
                    'Accuracy': accuracy.strip(),
                }
        
        return moves


class PokemonDataProcessor:
    def __init__(self, input_file, locations_file, moves_file, output_file, abilities_file, ignore_list, key_order, moves_tab, poke_info_tab, misc_tab):
        self.input_file = input_file
        self.locations_file = locations_file
        self.moves_file = moves_file
        self.abilities_file = abilities_file
        self.output_file = output_file
        self.ignore_list = ignore_list
        self.key_order = key_order
        self.moves_tab = moves_tab
        self.poke_info_tab = poke_info_tab
        self.misc_tab = misc_tab
        self.all_data = []
        self.location_data = {}
        self.move_parser = MoveParser(self.moves_file)
        self.abilities = self.retrieve_ability_descriptions()  # Load ability descriptions

    def retrieve_ability_descriptions(self):
        abilities = {}
        with open(self.abilities_file, 'r') as file:
            current_ability = None
            for line in file:
                line = line.strip()
                if line.startswith("[") and line.endswith("]"):  # Start of a new ability block
                    current_ability = line[1:-1].strip().lower()  # Use the identifier as the key, normalized to lowercase
                elif line.startswith("Description") and current_ability:
                    description = line.split('=')[1].strip().strip('"')
                    abilities[current_ability] = description  # Use the identifier as the key
                    current_ability = None  # Reset for the next block
        return abilities




    def process_input_file(self):
        with open(self.input_file, 'r') as file:
            lines = file.readlines()

        data = {}
        for line in lines:
            line = line.strip()
            if line == '#-------------------------------':
                if data:
                    if 'InternalName' not in data or data['InternalName'] is None:
                        data['InternalName'] = data['Name'].upper()
                    data['UniqueID'] = f"{data['Name']}_{data['InternalName']}"  # Generate UniqueID

                    self.process_abilities(data)

                    # Check if InternalName is EEVEE and update Evolutions field
                    if data.get('InternalName') == 'EEVEE':
                        if 'Evolutions' in data:
                            data['Evolutions'] += '(32)'
                        else:
                            data['Evolutions'] = '(32)'

                    self.all_data.append(data)
                    data = {}
            elif '=' in line:
                key, value = line.split('=', 1)
                key = key.strip()
                value = value.strip()
                if key in self.ignore_list and key != 'Types':
                    continue
                if key == 'BaseStats':
                    stats = value.split(',')
                    data['HP'] = stats[0]
                    data['Attack'] = stats[1]
                    data['Defense'] = stats[2]
                    data['Spd'] = stats[3]
                    data['SpAtk'] = stats[4]
                    data['SpDef'] = stats[5]
                elif key == 'HiddenAbility':
                    abilities = value.split(',')
                    for ability in abilities:
                        data['HiddenAbility'] = ability
                elif key == 'Types':
                    types = value.split(',')
                    data['Type1'] = types[0]
                    if len(types) > 1:
                        data['Type2'] = types[1]
                    else:
                        data['Type2'] = None
                else:
                    data[key] = value

        if data:
            if 'InternalName' not in data or data['InternalName'] is None:
                data['InternalName'] = data['Name'].upper()
            data['UniqueID'] = f"{data['Name']}_{data['InternalName']}"

            self.process_abilities(data)

            # Check if InternalName is EEVEE and update Evolutions field
            if data.get('InternalName') == 'EEVEE':
                if 'Evolutions' in data:
                    data['Evolutions'] += '(32)'
                else:
                    data['Evolutions'] = '(32)'

            self.all_data.append(data)

    def process_abilities(self, data):
        for column in ['Abilities', 'HiddenAbility']:
            if column in data:
                abilities_column = data[column]
                if abilities_column:
                    abilities = abilities_column.split(',')
                    formatted_abilities = []
                    for ability in abilities:
                        ability_name = ability.strip().lower()  # Normalize to lowercase
                        if ability_name in self.abilities:  # Check against identifiers, not names
                            description = self.abilities[ability_name]
                            formatted_ability = f"{ability_name.capitalize()} - \"{description}\""  # Capitalize for display
                            formatted_abilities.append(formatted_ability)
                    data[column] = ', '.join(formatted_abilities)


          
    # Split and process encounters file into Python data dictionary 
    def process_locations_file(self):
        with open(self.locations_file, 'r') as file:
            lines = file.readlines()
        
        current_location = None
        for line in lines:
            line = line.strip()
            if line.startswith('#'):
                continue
            if line.startswith('['):
                current_location = line.split('#')[1].strip()
            elif ',' in line:
                parts = line.split(',')
                if len(parts) == 4:
                    _, name, level_start, level_end = parts
                    level_range = f"{level_start}-{level_end}"
                    location_entry = f"{current_location}: {level_range}"
                    if name not in self.location_data:
                        self.location_data[name] = []
                    self.location_data[name].append(location_entry)

    # Add location data to pokemon names
    def add_location_data(self):
        for entry in self.all_data:
            name = entry.get('InternalName', '')
            if name in self.location_data:
                entry['Location Found'] = "; ".join(self.location_data[name])
            else:
                entry['Location Found'] = ''
        
        if 'Location Found' not in self.key_order:
            self.key_order.append('Location Found')

    # Update the move columns with parsed move details
    def update_moves_columns(self, df):
        def update_column(cell, column_type):
            if pd.isna(cell):
                return cell
            
            moves = cell.split(',')
            updated_moves = []
            
            if column_type == 'Moves':
                i = 0
                while i < len(moves):
                    level = moves[i].strip()
                    move = moves[i + 1].strip().upper()
                    if move == '':
                        i += 2
                        continue  # Skip to the next iteration of the loop
                    
                    if move in self.move_parser.moves:
                        data = self.move_parser.moves[move]
                        updated_move = f"{level} - {data['Name']} - Type: {data['Type']} - Power: {data['Power']} - Accuracy: {data['Accuracy']} - Category: {data['Category']}"
                        updated_moves.append(updated_move)
                    
                    i += 2
            
            elif column_type in ['TutorMoves', 'EggMoves']:
                for move in moves:
                    move = move.strip().upper()
                    if move == '':
                        continue  # Skip empty moves
                    
                    if move in self.move_parser.moves:
                        data = self.move_parser.moves[move]
                        updated_move = f"{data['Name']} - Type: {data['Type']} - Power: {data['Power']} - Accuracy: {data['Accuracy']} - Category: {data['Category']}"
                        updated_moves.append(updated_move)
            
            return ', '.join(updated_moves)

        for column, column_type in [('Moves', 'Moves'), ('TutorMoves', 'TutorMoves'), ('EggMoves', 'EggMoves')]:
            if column in df.columns:
                df[column] = df[column].apply(update_column, column_type=column_type)
        
        return df

    # Build Excel workbook
    def create_workbook(self):
        wb = Workbook()

        def add_data_to_sheet(sheet_name, data, headers):
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)
            for entry in data:
                row = [entry.get(header, '') for header in headers]
                ws.append(row)

        main_headers = ['InternalName', 'Name'] + self.key_order if 'InternalName' not in self.key_order else self.key_order
        add_data_to_sheet("Main", self.all_data, main_headers)
        add_data_to_sheet("Poke Info", self.all_data, ['InternalName', 'Name'] + self.poke_info_tab)
        add_data_to_sheet("Misc", self.all_data, ['InternalName', 'Name'] + self.misc_tab)

        moves_headers = ['InternalName', 'Moves', 'TutorMoves', 'EggMoves']
        moves_data = [{k: entry.get(k, '') for k in moves_headers} for entry in self.all_data]

        df_moves = pd.DataFrame(moves_data)
        df_moves = self.update_moves_columns(df_moves)

        with pd.ExcelWriter(self.output_file, engine='openpyxl', mode='w') as writer:
            writer.book = wb
            df_moves.to_excel(writer, sheet_name='Moves', index=False)

        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        wb.save(self.output_file)


    # Check if the Excel sheet exists and has all required columns
    def check_excel_sheet(self):
        if not os.path.exists(self.output_file):
            return False
        
        required_columns = ['Name', 'InternalName', 'FormName', 'Type1', 'Type2', 'HP', 'Attack', 'Defense', 'SpAtk', 'SpDef', 'Spd', 'Abilities', 'HiddenAbility', 'Evolutions', 'Location Found', 'Evolution Line']
        
        try:
            df = pd.read_excel(self.output_file, sheet_name="Main")
            columns = df.columns.tolist()
            
            if set(required_columns).issubset(set(columns)):
                return True
            else:
                return False
                
        except Exception as e:
            return False
    
    # Check if the Excel sheet was created in the last 2 days
    def check_excel_creation_date(self):
        if os.path.exists(self.output_file):
            creation_time = datetime.datetime.fromtimestamp(os.path.getctime(self.output_file))
            current_time = datetime.datetime.now()
            if (current_time - creation_time).days <= 2:
                return True
        return False

    # Run all functions
    def run(self):
        #if self.check_excel_sheet() and self.check_excel_creation_date():
            #pass
        #else:
            self.process_input_file()
            self.process_locations_file()
            self.add_location_data()
            self.create_workbook()

base_dir = os.path.dirname(os.path.abspath(__file__))

# Paths to the files
workbook_path = os.path.join('PBS', 'pokemon.txt')
encounter_path = os.path.join('PBS', 'encounters.txt')
moves_path = os.path.join('PBS', 'moves.txt')
abilities_path = os.path.join('PBS', 'abilities.txt')

# Configuration
input_file = workbook_path
locations_file = encounter_path
moves_file = moves_path
abilities_file = abilities_path
output_file = 'pokedexCEL.xlsx'
ignore_list = ['BaseExp', 'BattlerEnemyX', 'BattlerEnemyY', 'BattlerPlayerX', 'BattlerPlayerY', 'BattlerShadowSize', 'BattlerShadowX', 'CatchRate', 'Types', 'Category', 'EggGroups', 'HatchSteps', 'EVs', 'EffortPoints', 'HiddenAbilities']
key_order = ['UniqueID', 'Name', 'InternalName', 'FormName', 'Type1', 'Type2', 'HP', 'Attack', 'Defense', 'SpAtk', 'SpDef', 'Spd', 'Abilities', 'HiddenAbility', 'Evolutions', 'Location Found']
moves_tab = ['UniqueID', 'InternalName', 'Moves', 'TutorMoves', 'EggMoves']
poke_info_tab = ['UniqueID', 'Generation', 'Height', 'Weight', 'Color', 'Pokedex', 'Shape', 'Kind', 'GrowthRate', 'GenderRate', 'GenderRatio', 'Habitat', 'BaseEXP', 'Rareness', 'Happiness', 'Compatibility', 'Incense']
misc_tab = ['WildItemCommon', 'WildItemRare', 'WildItemUncommon', 'StepsToHatch']

# Run the processor
processor = PokemonDataProcessor(input_file, locations_file, moves_file, output_file, abilities_file, ignore_list, key_order, moves_tab, poke_info_tab, misc_tab)

def main():
    processor.run()

#if __name__ == "__main__":
    #main()
    
    
